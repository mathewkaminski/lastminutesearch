import sys
import tempfile
import os
from extract_financials import check_deps


def test_missing_dependency_message():
    """check_deps() prints install instructions for missing packages."""
    missing = check_deps(required=["nonexistent_package_xyz"])
    assert len(missing) == 1
    assert "nonexistent_package_xyz" in missing[0]

def test_no_missing_dependencies():
    """check_deps() returns empty list when all packages present."""
    missing = check_deps(required=["sys", "os"])
    assert missing == []


def test_extract_all_pages_returns_list():
    """extract_pages() returns a list of dicts with page_number, text, tables."""
    from extract_financials import extract_pages
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    assert isinstance(pages, list)
    assert len(pages) > 0
    assert "page_number" in pages[0]
    assert "text" in pages[0]
    assert "tables" in pages[0]

def test_extract_pages_includes_text():
    """Pages with known content return non-empty text."""
    from extract_financials import extract_pages
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    # Page 8 (index 7) has Revenue data in it
    page_8 = next(p for p in pages if p["page_number"] == 8)
    assert "Revenue" in page_8["text"]


def test_detect_financial_pages_returns_subset():
    """detect_financial_pages() returns only pages with financial data."""
    from extract_financials import extract_pages, detect_financial_pages
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    financial = detect_financial_pages(pages)
    assert isinstance(financial, list)
    # JAM deck has at least 4 financial pages
    assert len(financial) >= 4
    # Each result has required fields
    assert all("page_number" in f for f in financial)
    assert all("tab_name" in f for f in financial)

def test_detect_financial_pages_excludes_non_financial():
    """detect_financial_pages() does not include text-only slides."""
    from extract_financials import extract_pages, detect_financial_pages
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    financial = detect_financial_pages(pages)
    page_numbers = [f["page_number"] for f in financial]
    # Page 1 is a cover slide — should not appear
    assert 1 not in page_numbers


def test_extract_page_data_structure():
    """extract_page_data() returns tab_name, table rows, and notes list."""
    from extract_financials import extract_pages, extract_page_data
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    page_8 = next(p for p in pages if p["page_number"] == 8)
    result = extract_page_data(page_8, tab_name="Financials 2022-2025")
    assert result["tab_name"] == "Financials 2022-2025"
    assert isinstance(result["table"], list)
    assert len(result["table"]) > 0
    assert isinstance(result["notes"], list)

def test_extract_page_data_strips_currency_symbols():
    """Numeric values like $9,470,051 are returned as plain numbers (9470051)."""
    from extract_financials import extract_pages, extract_page_data
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    page_8 = next(p for p in pages if p["page_number"] == 8)
    result = extract_page_data(page_8, tab_name="Financials 2022-2025")
    all_values = [cell for row in result["table"] for cell in row]
    assert any("9470051" in str(v) for v in all_values)
    assert not any("$" in str(v) for v in all_values)

def test_extract_page_data_captures_notes():
    """Notes/bullet text is captured even when outside table boundaries."""
    from extract_financials import extract_pages, extract_page_data
    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    pages = extract_pages(pdf_path)
    page_8 = next(p for p in pages if p["page_number"] == 8)
    result = extract_page_data(page_8, tab_name="Financials 2022-2025")
    combined_notes = " ".join(result["notes"])
    assert "fiscal year" in combined_notes.lower() or "August" in combined_notes


def test_write_excel_creates_file():
    """write_excel() creates an xlsx file at the given path."""
    from extract_financials import write_excel
    import openpyxl

    pages_data = [
        {
            "tab_name": "Test Financials",
            "table": [["Year", "Revenue"], ["2024", "$1,000,000"]],
            "notes": ["This is a test note."]
        }
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "test_output.xlsx")
        write_excel(pages_data, out_path)
        assert os.path.exists(out_path)
        wb = openpyxl.load_workbook(out_path)
        assert "Test Financials" in wb.sheetnames

def test_write_excel_table_content():
    """write_excel() writes table rows starting at A1."""
    from extract_financials import write_excel
    import openpyxl

    pages_data = [
        {
            "tab_name": "Revenue",
            "table": [["Region", "2024"], ["Toronto", "$8,203,397"]],
            "notes": ["Toronto is largest market."]
        }
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "test.xlsx")
        write_excel(pages_data, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Revenue"]
        assert ws["A1"].value == "Region"
        assert ws["B2"].value == "$8,203,397"

def test_write_excel_notes_below_table():
    """Notes appear below table with a blank row separator."""
    from extract_financials import write_excel
    import openpyxl

    pages_data = [
        {
            "tab_name": "Sheet1",
            "table": [["Col1"], ["Val1"]],
            "notes": ["First note.", "Second note."]
        }
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "test.xlsx")
        write_excel(pages_data, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Sheet1"]
        # Row 1: header, Row 2: data, Row 3: blank, Row 4: "Notes:", Row 5+: notes
        assert ws["A4"].value == "Notes:"
        assert ws["A5"].value == "First note."
        assert ws["A6"].value == "Second note."


def test_full_pipeline_jam_deck():
    """End-to-end: JAM PDF produces xlsx with expected tabs."""
    import subprocess
    import openpyxl

    pdf_path = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin.pdf"
    expected_xlsx = r"C:\Users\mathe\OneDrive\Documents\Consulting\JAM\JAM - Investor Deck for Mat & Orin_financials.xlsx"

    # Clean up any prior output
    if os.path.exists(expected_xlsx):
        os.remove(expected_xlsx)

    # Load API key into environment for subprocess
    from pathlib import Path
    env = os.environ.copy()
    for line in Path(r"C:\Users\mathe\VSCode\aa_RecLeagueDB\.env").read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()

    result = subprocess.run(
        ["python", "extract_financials.py", pdf_path],
        capture_output=True, text=True, env=env,
        cwd=r"C:\Users\mathe\.claude\skills\extract-financials"
    )
    assert result.returncode == 0, f"Script failed:\n{result.stderr}\n{result.stdout}"
    assert os.path.exists(expected_xlsx), "Output file not created"

    wb = openpyxl.load_workbook(expected_xlsx)
    assert len(wb.sheetnames) >= 4, f"Expected >= 4 sheets, got: {wb.sheetnames}"
    print(f"Created sheets: {wb.sheetnames}")
